from flask import Flask, render_template, request, redirect, send_file, jsonify
from openpyxl import load_workbook, Workbook
import os
import psycopg2
from psycopg2 import pool
from urllib.parse import urlparse
import uuid

db_initialized = False

app = Flask(__name__)

UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

db_pool = None


# ---------------- DB POOL ---------------- #
def init_pool():
    global db_pool

    database_url = os.environ.get("DATABASE_URL")
    if not database_url:
        raise Exception("DATABASE_URL not set")

    result = urlparse(database_url)

    db_pool = psycopg2.pool.SimpleConnectionPool(
        1, 10,
        database=result.path[1:],
        user=result.username,
        password=result.password,
        host=result.hostname,
        port=result.port,
        sslmode="require",
        connect_timeout=5
    )



def get_connection():
    global db_pool, db_initialized

    try:
        # 🔵 ยังไม่มี pool → สร้าง
        if db_pool is None:
            print("🔵 INIT POOL")
            init_pool()

        # 🟢 ยังไม่ init DB → สร้าง table
        if not db_initialized:
            print("🟢 INIT DB")
            init_db()   # ❗ ไม่ต้อง set db_initialized ตรงนี้

        # 📦 ดึง connection
        conn = db_pool.getconn()

        # ✅ เช็คว่า connection ยังไม่ตาย
        cur = conn.cursor()
        cur.execute("SELECT 1")
        cur.close()

        return conn

    except Exception as e:
        print("❌ DB ERROR → reconnect:", e)

        # 🔥 ปิด pool เก่าทิ้ง (กันค้างสะสม)
        try:
            if db_pool:
                db_pool.closeall()
        except:
            pass

        # 🔁 สร้างใหม่
        init_pool()

        # 📦 เอา connection ใหม่
        conn = db_pool.getconn()
        return conn




def release_connection(conn):
    if db_pool:
        db_pool.putconn(conn)


# ---------------- INIT DB ---------------- #
def init_db():
    global db_initialized

    if db_initialized:
        return

    conn = db_pool.getconn()
    try:
        cur = conn.cursor()

        # warehouses
        cur.execute("""
        CREATE TABLE IF NOT EXISTS warehouses (
            id SERIAL PRIMARY KEY,
            name TEXT UNIQUE
        )
        """)

        # products
        cur.execute("""
        CREATE TABLE IF NOT EXISTS products (
            id SERIAL PRIMARY KEY,
            warehouse TEXT,
            location TEXT,
            model TEXT,
            description TEXT,
            inv_qty INTEGER DEFAULT 0,
            act_qty INTEGER DEFAULT 0
        )
        """)

        # scans
        cur.execute("""
        CREATE TABLE IF NOT EXISTS scans (
            id SERIAL PRIMARY KEY,
            full_barcode TEXT,
            warehouse TEXT,
            UNIQUE(full_barcode, warehouse)
        )
        """)

        # index
        cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_products_model_wh
        ON products(model, warehouse)
        """)

        conn.commit()
        cur.close()

        db_initialized = True

    except Exception as e:
        conn.rollback()
        print("INIT DB ERROR:", e)

    finally:
        db_pool.putconn(conn)




# ---------------- CACHE ---------------- #
@app.after_request
def add_header(response):
    response.headers["Cache-Control"] = "no-store"
    return response


# ---------------- FUNCTIONS ---------------- #
def get_warehouses():
    conn = get_connection()
    try:
        c = conn.cursor()
        c.execute("SELECT name FROM warehouses ORDER BY name")
        return [r[0] for r in c.fetchall()]
    except Exception as e:
        print("GET WAREHOUSE ERROR:", e)
        return []
    finally:
        release_connection(conn)


def get_products(warehouse):
    conn = get_connection()
    try:
        c = conn.cursor()
        c.execute("""
            SELECT id, location, model, description, inv_qty, act_qty
            FROM products
            WHERE warehouse=%s
        """, (warehouse,))
        return c.fetchall()
    except Exception as e:
        print("GET PRODUCT ERROR:", e)
        return []
    finally:
        release_connection(conn)


# ---------------- ROUTES ---------------- #
@app.route("/")
def index():
    warehouses = get_warehouses()
    if not warehouses:
        return "OK"
    return redirect(f"/warehouse/{warehouses[0]}")


@app.route("/add_warehouse", methods=["POST"])
def add_warehouse():
    data = request.get_json()
    name = data.get("name")

    if not name:
        return jsonify({"success": False})

    conn = get_connection()
    try:
        c = conn.cursor()
        c.execute("INSERT INTO warehouses (name) VALUES (%s)", (name,))
        conn.commit()
        return jsonify({"success": True})
    except:
        conn.rollback()
        return jsonify({"success": False})
    finally:
        release_connection(conn)


@app.route("/warehouse/<warehouse>")
def warehouse_page(warehouse):
    warehouses = get_warehouses()
    products = get_products(warehouse)

    return render_template(
        "warehouse.html",
        warehouses=warehouses,
        warehouse=warehouse,
        products=products
    )


# ---------------- IMPORT ---------------- #
@app.route("/import/<warehouse>", methods=["POST"])
def import_excel(warehouse):
    conn = get_connection()
    try:
        file = request.files["file"]

        # ✅ กันชื่อไฟล์ซ้ำ
        filename = str(uuid.uuid4()) + ".xlsx"
        path = os.path.join(UPLOAD_FOLDER, filename)
        file.save(path)

        wb = load_workbook(path)
        ws = wb.active

        c = conn.cursor()

        # ✅ ลบแล้ว insert ใน transaction เดียว (กันพัง)
        c.execute("DELETE FROM products WHERE warehouse=%s", (warehouse,))
        c.execute("DELETE FROM scans WHERE warehouse=%s", (warehouse,))

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[1]:
                continue

            location, model, description, inv_qty = row[:4]

            c.execute("""
                INSERT INTO products
                (warehouse, location, model, description, inv_qty, act_qty)
                VALUES (%s,%s,%s,%s,%s,0)
            """, (warehouse, location, model, description, inv_qty or 0))

        conn.commit()
        return redirect(f"/warehouse/{warehouse}")

    except Exception as e:
        conn.rollback()
        print("IMPORT ERROR:", e)
        return "IMPORT FAIL"

    finally:
        release_connection(conn)


# ---------------- SCAN ---------------- #
@app.route("/scan", methods=["POST"])
def scan():
    conn = get_connection()
    try:
        barcode = request.form.get("barcode")
        warehouse = request.form.get("warehouse")

        if not barcode:
            return jsonify({"status": "not_found"})

        model = barcode[:9].upper()

        cur = conn.cursor()

        cur.execute("""
            SELECT id
            FROM products
            WHERE model=%s AND warehouse=%s
        """, (model, warehouse))

        row = cur.fetchone()

        if not row:
            return jsonify({"status": "not_found"})

        product_id = row[0]

        # ✅ กัน scan ซ้ำ
        cur.execute("""
            INSERT INTO scans (full_barcode, warehouse)
            VALUES (%s,%s)
            ON CONFLICT DO NOTHING
        """, (barcode, warehouse))

        if cur.rowcount == 0:
            return jsonify({"status": "duplicate"})

        # ✅ FIX สำคัญ: กันค่าพัง
        cur.execute("""
            UPDATE products
            SET act_qty = act_qty + 1
            WHERE id=%s
        """, (product_id,))

        conn.commit()
        return jsonify({"status": "success"})

    except Exception as e:
        conn.rollback()
        print("SCAN ERROR:", e)
        return jsonify({"status": "error"})

    finally:
        release_connection(conn)

# ---------------- New Barcode ---------------- #
@app.route("/add_new_barcode", methods=["POST"])
def add_new_barcode():
    conn = get_connection()
    try:
        data = request.get_json()

        barcode = data.get("barcode")
        warehouse = data.get("warehouse")
        location = data.get("location")

        if not barcode:
            return jsonify({"success": False})

        model = barcode[:9].upper()

        cur = conn.cursor()

        # 🔍 เช็คว่ามีอยู่แล้วไหม
        cur.execute("""
            SELECT id, inv_qty
            FROM products
            WHERE model=%s AND warehouse=%s
        """, (model, warehouse))

        row = cur.fetchone()

        if row:
            # 👉 มีอยู่แล้ว → บวกเพิ่ม
            product_id, inv_qty = row

            new_qty = (inv_qty or 0) + 1

            cur.execute("""
                UPDATE products
                SET inv_qty=%s,
                    act_qty=%s
                WHERE id=%s
            """, (new_qty, new_qty, product_id))

        else:
            # 👉 ยังไม่มี → สร้างใหม่
            cur.execute("""
                INSERT INTO products
                (warehouse, location, model, description, inv_qty, act_qty)
                VALUES (%s,%s,%s,%s,1,1)
            """, (
                warehouse,
                location,
                model,
                "ไม่มีในฐานข้อมูล"
            ))

        conn.commit()
        return jsonify({"success": True})

    except Exception as e:
        conn.rollback()
        print("ADD NEW ERROR:", e)
        return jsonify({"success": False})

    finally:
        release_connection(conn)


# ---------------- DELETE ---------------- #
@app.route("/delete_selected", methods=["POST"])
def delete_selected():
    conn = get_connection()
    try:
        data = request.get_json()
        ids = data.get("ids", [])

        if not ids:
            return "No data"

        ids = [int(i) for i in ids]

        c = conn.cursor()
        c.execute(
            "DELETE FROM products WHERE id = ANY(%s::int[])",
            (ids,)
        )

        conn.commit()
        return "OK"

    except Exception as e:
        conn.rollback()
        print("DELETE ERROR:", e)
        return "ERROR"

    finally:
        release_connection(conn)


# ---------------- EXPORT ---------------- #
@app.route("/export/<warehouse>")
def export_excel(warehouse):
    conn = get_connection()
    try:
        c = conn.cursor()

        c.execute("""
            SELECT warehouse, location, model, description, inv_qty, act_qty
            FROM products WHERE warehouse=%s
        """, (warehouse,))


        rows = c.fetchall()

        wb = Workbook()
        ws = wb.active

        # 🔵 หัวรายงาน
        ws["A1"] = "Warehouse"
        ws["B1"] = warehouse

        # 🔵 เว้นบรรทัด
        ws.append([])

        # 🟢 หัวตาราง
        ws.append(["Location", "Model Code", "Product description", "Inv.Qty", "Act.Qty"])

        # 📦 ข้อมูล
        for row in rows:
            ws.append(row)


        file_path = os.path.join(UPLOAD_FOLDER, f"{warehouse}_result.xlsx")
        wb.save(file_path)

        return send_file(file_path, as_attachment=True)

    except Exception as e:
        conn.rollback()
        print("EXPORT ERROR:", e)
        return "EXPORT FAIL"

    finally:
        release_connection(conn)


# ---------------- RUN ---------------- #
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)
